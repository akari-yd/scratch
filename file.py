import openpyxl
import re
import requests


# name:'.xlsx'文件名，不存在则生成表格，返回值：表格
def getxlsx(name):
    try:
        excel = openpyxl.load_workbook(name)
        return excel
    except:
        return 0


# excel:表格,sort_condition:排序依据,derect:排序方向0/1（升序/降序)
def sortxlsx(name, sort_condition, direct=0):
    excel = getxlsx(name)
    if excel:
        raw_table = excel.active
        aim_table = excel.create_sheet(sort_condition)
        rows = list(raw_table.rows)
        # max_rows = raw_table.max_row
        first_row = []
        for cell in rows[0]:
            first_row.append(cell.value)
        aim_table.append(first_row)
        column = 0
        for tag in rows[0]:
            if tag.value == sort_condition:
                count = 0
                for number in range(1, len(rows)):
                    count = count + 1
                    aim_row = rows[1]
                    for row in rows[1:]:
                        if direct == 0:
                            if row[column].value < aim_row[column].value:
                                aim_row = row
                        else:
                            if row[column].value > aim_row[column].value:
                                aim_row = row
                    aim = [number]
                    for cell in aim_row[1:]:
                        aim.append(cell.value)
                    rows.remove(aim_row)
                    aim_table.append(aim)
            column = column + 1
        excel.save(name)


# name:file name   content:the content of the file
def filesave(name, content):
    table = openpyxl.Workbook()
    sheet = table.active
    for row in content:
        sheet.append(row)
    table.save(name)
